"""
email_parser.py
네이버 이메일에서 은행/카드 알림을 파싱해 Google Sheets에 저장하고,
파싱 대상 + 광고/쇼핑/뉴스레터/SNS 메일을 자동으로 폴더 이동합니다.
"""

import base64
import email
import html as html_module
import imaplib
import os
import re
from datetime import datetime, timedelta
from email.header import decode_header
from email.utils import parsedate_to_datetime

# ── 환경설정 ──────────────────────────────────────────
NAVER_EMAIL = os.environ.get("NAVER_EMAIL", "")
NAVER_APP_PW = os.environ.get("NAVER_APP_PW", "")
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "")
GOOGLE_CREDS_JSON = os.environ.get("GOOGLE_CREDS_JSON", "")
ENABLE_EMAIL_CLEANUP = os.environ.get("ENABLE_EMAIL_CLEANUP", "").lower() in ("1", "true", "yes")
LOOKBACK_HOURS = int(os.environ.get("LOOKBACK_HOURS", "2"))
# BC카드 명세서 PDF 비밀번호 (생년월일 6자리). 미설정 시 PDF 파싱 건너뜀.
BC_PDF_PASSWORD = os.environ.get("BC_PDF_PASSWORD", "")
# 카카오뱅크 '거래내역 엑셀' 첨부 비밀번호 (생년월일 6자리).
# 앱 → 거래내역 → 내보내기(이메일)로 받은 암호화 xlsx를 자동 파싱할 때 사용.
KAKAO_XLSX_PASSWORD = os.environ.get("KAKAO_XLSX_PASSWORD", "")
# 명세서를 한 번 파싱한 후에도 다음 달까지 LOOKBACK_HOURS 윈도우 밖에 머물 수 있으므로
# 명세서 메일 검색은 별도로 더 긴 윈도우(기본 35일)를 사용한다.
STATEMENT_LOOKBACK_DAYS = int(os.environ.get("STATEMENT_LOOKBACK_DAYS", "35"))

IMAP_HOST = "imap.naver.com"
IMAP_PORT = 993

# 분류된 이메일이 이동할 폴더 (UTF-8, IMAP UTF-7로 자동 인코딩)
PROCESSED_FOLDER = "가계부_처리완료"
AD_FOLDER = "가계부_광고"
SHOPPING_FOLDER = "가계부_쇼핑"
NEWSLETTER_FOLDER = "가계부_뉴스레터"
SNS_FOLDER = "가계부_SNS"
STATEMENT_FOLDER = "가계부_명세서"  # PDF 파싱 실패 시 보관용

# ── 발신자 → 출처 매핑 (결제/은행 알림만) ─────────────
SENDER_PATTERNS = {
    "BC카드": ["bcbill@bccard.com", "bccard.com"],
    "카카오뱅크": ["no-reply@mail.kakaobank.com", "kakaobank.com"],
    "현대카드": ["admin@hyundaicard.com", "hyundaicard.com"],
    "KB카드": ["kbcard.com", "kbcard.co.kr"],  # parse_kb_email_html + process_kb_statements 처리
    "IBK기업은행": ["ibk.co.kr"],
    "네이버페이": ["naverpayadmin_noreply@navercorp.com"],
    "토스페이먼츠": ["bill@bill-mail.tosspayments.com", "tosspayments.com"],
    "나이스정보통신": ["nice_customer@nicepg.co.kr"],
    "헥토파이낸셜": ["noreply@hecto.co.kr"],
}

# 파싱 대상 이메일이 들어 있을 수 있는 폴더 (검색 대상)
IMAP_FOLDERS = [
    "INBOX",
    "&zK2tbAC3rLDIHA-",   # 청구/결제 폴더
    "&yPy7OA-|&vDDBoQ-",  # 네이버페이 등
]

# ── 비거래 이메일 분류 규칙 ───────────────────────────
NON_TX_CATEGORIES = [
    # (카테고리명, 폴더, sender 키워드, subject 키워드)
    # subject 매칭은 classify_non_transaction에서 공백 제거 후 substring 검사
    ("쇼핑", SHOPPING_FOLDER,
     ["coupang.com", "11st.co.kr", "ssg.com", "gmarket", "auction.co.kr",
      "wemakeprice", "smartstore.naver.com", "ohou.se", "kurly.com",
      "musinsa", "29cm",
      "aliexpress.com", "amazon.com", "amazon.co.jp", "qoo10",
      "tmall", "taobao", "shopee", "lotteon",
      # PG사·구독·고지서·가맹점 영수증 (카드 알림은 SENDER_PATTERNS가 먼저 잡음)
      "easypay.co.kr", "kcp.co.kr", "apti.co.kr",
      "tving.com", "email.apple.com",
      "gayamy.co.kr"],
     ["주문확인", "주문완료", "주문건", "주문내역",
      "배송완료", "배송시작", "배송중", "배송재개",
      "발송완료", "출고완료", "도착예정", "도착",
      "배송지연", "반품접수", "교환접수", "구매확정",
      "운송장", "택배사", "송장번호",
      # PG/영수증/고지서/정기결제 패턴 (카드 알림은 위에서 이미 처리됨)
      "결제하신", "결제내역", "결제 내역", "정기결제",
      "결제 예정", "영수증", "고지서"]),
    ("SNS", SNS_FOLDER,
     ["instagram.com", "facebookmail.com", "linkedin.com", "twitter.com",
      "x.com", "tiktok.com", "youtube.com", "discord.com", "slack.com",
      "github.com"],
     []),
    ("뉴스레터", NEWSLETTER_FOLDER,
     ["substack.com", "mailchimp", "mailchi.mp", "newsletter@",
      "@news.", ".letter", "stibee.com", "maily.so"],
     ["뉴스레터", "newsletter", "주간", "월간", "weekly digest", "daily digest"]),
    ("광고", AD_FOLDER,
     [],
     ["(광고)", "[광고]", "광고)", "이벤트", "쿠폰", "할인", "프로모션",
      "특가", "혜택", "당첨", "추첨", "기획전", "세일", "초대권"]),
]

# ── 금액/유형 키워드 ──────────────────────────────────
# "1,000원 결제" / "결제 1,000원" 형태 우선, 그 다음 fallback
AMOUNT_NEAR_KEYWORD = re.compile(
    r"(?:(출금|이체|결제|승인|사용|입금|수신|환불|취소)\s*[:\s]?\s*([0-9,]+)\s*원"
    r"|([0-9,]+)\s*원\s*(?:이?\s*)?(출금|이체|결제|승인|사용|입금|수신|환불|취소))"
)
AMOUNT_FALLBACK = re.compile(r"([0-9,]+)\s*원")

# 무시해야 하는 컨텍스트 (잔액/한도/포인트 등)
AMOUNT_BLACKLIST_CONTEXT = [
    "잔액", "한도", "포인트", "마일리지", "적립", "누적", "총액", "현재잔액",
    "사용가능", "이용가능", "혜택받은", "할인받은", "사용한도", "이용한도",
]

INCOME_KEYWORDS = ["입금", "수신", "급여", "이자", "환급", "환불", "취소"]
EXPENSE_KEYWORDS = ["출금", "이체", "결제", "승인", "사용"]

CATEGORY_KEYWORDS = {
    "식비": ["식당", "음식", "카페", "커피", "배달", "맥도날드", "스타벅스", "버거킹",
             "편의점", "GS25", "CU", "세븐", "이마트24", "투썸", "메가", "공차",
             "BBQ", "교촌", "도미노", "피자", "짬뽕", "국수", "순대", "칼국수",
             "수산", "분식", "치킨", "족발"],
    "교통": ["택시", "버스", "지하철", "주유", "카카오택시", "티머니", "하이패스",
             "S-OIL", "SK에너지", "GS칼텍스", "현대오일뱅크", "철도", "코레일"],
    "쇼핑": ["쿠팡", "네이버", "G마켓", "옥션", "11번가", "이마트", "홈플러스",
             "코스트코", "마켓컬리", "올리브영", "다이소", "무신사"],
    "의료": ["병원", "약국", "의원", "클리닉", "치과", "한의원"],
    "통신": ["SKT", "KT", "LG", "통신", "인터넷", "헬로비전"],
    "구독": ["넷플릭스", "유튜브", "스포티파이", "왓챠", "애플", "MS", "어도비",
             "디즈니", "티빙", "웨이브"],
    "주거": ["관리비", "전기", "수도", "가스", "월세", "임대료", "한국전력", "도시가스"],
    "문화": ["CGV", "메가박스", "롯데시네마", "영화", "공연", "박물관", "전시"],
    "미용": ["헤어", "미용실", "이발", "네일", "피부관리", "뷰티"],
    "수입": ["급여", "이자", "환급", "월급", "보너스"],
}


# ── IMAP UTF-7 인코딩 (RFC 3501 modified UTF-7) ──────
def imap_utf7_encode(s: str) -> str:
    """폴더명을 IMAP modified UTF-7로 인코딩"""
    result = []
    buf = []

    def flush():
        if buf:
            encoded = base64.b64encode("".join(buf).encode("utf-16-be")).rstrip(b"=").decode("ascii")
            result.append("&" + encoded.replace("/", ",") + "-")
            buf.clear()

    for ch in s:
        code = ord(ch)
        if 0x20 <= code <= 0x7E and ch != "&":
            flush()
            result.append(ch)
        elif ch == "&":
            flush()
            result.append("&-")
        else:
            buf.append(ch)
    flush()
    return "".join(result)


# ── 헤더/본문 파싱 ────────────────────────────────────
def decode_str(s):
    if s is None:
        return ""
    parts = decode_header(s)
    out = ""
    for part, charset in parts:
        if isinstance(part, bytes):
            cs = charset or "utf-8"
            try:
                out += part.decode(cs, errors="replace")
            except (LookupError, UnicodeDecodeError):
                out += part.decode("utf-8", errors="replace")
        elif isinstance(part, str):
            out += part
        else:
            # 일부 메일은 decode_header가 email.header.Header를 반환할 수 있음
            out += str(part)
    return out


def header_str(value) -> str:
    """msg.get(...) 결과를 안전하게 str로. None/Header/str 모두 처리."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def strip_html(html: str) -> str:
    """HTML 태그 제거 + 엔티티 디코딩 + 공백 정리"""
    # <script>, <style> 블록 통째로 제거
    html = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", html, flags=re.IGNORECASE | re.DOTALL)
    # <br>, </p>, </div> 등 블록 끊기를 줄바꿈으로
    html = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
    html = re.sub(r"</(p|div|tr|li|td|h[1-6])>", "\n", html, flags=re.IGNORECASE)
    # 나머지 태그 제거
    html = re.sub(r"<[^>]+>", " ", html)
    # 엔티티 디코딩 (&nbsp; 등)
    html = html_module.unescape(html)
    # 공백/줄바꿈 정리
    html = re.sub(r"[ \t ]+", " ", html)
    html = re.sub(r"\n\s*\n+", "\n", html)
    return html.strip()


def get_email_body(msg) -> str:
    """text/plain을 우선, 없으면 text/html을 strip해서 반환"""
    plain = ""
    html = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_maintype() == "multipart":
                continue
            content_type = part.get_content_type()
            disposition = header_str(part.get("Content-Disposition")).lower()
            if "attachment" in disposition:
                continue
            try:
                payload = part.get_payload(decode=True)
                if payload is None:
                    continue
                charset = part.get_content_charset() or "utf-8"
                text = payload.decode(charset, errors="replace")
            except (LookupError, UnicodeDecodeError):
                continue
            if content_type == "text/plain":
                plain += text + "\n"
            elif content_type == "text/html":
                html += text + "\n"
    else:
        try:
            payload = msg.get_payload(decode=True)
            charset = msg.get_content_charset() or "utf-8"
            text = (payload or b"").decode(charset, errors="replace")
            if msg.get_content_type() == "text/html":
                html = text
            else:
                plain = text
        except (LookupError, UnicodeDecodeError):
            pass

    if plain.strip():
        return plain
    return strip_html(html)


# ── 금액/유형/가맹점 파싱 ─────────────────────────────
def _is_blacklisted_context(text: str, pos: int, window: int = 12) -> bool:
    """매치 위치 주변에 잔액/한도 등이 있는지"""
    start = max(0, pos - window)
    end = min(len(text), pos + window)
    snippet = text[start:end]
    return any(kw in snippet for kw in AMOUNT_BLACKLIST_CONTEXT)


def parse_amount(text: str) -> int | None:
    """본문에서 거래 금액 추출. 거래 키워드 근접 매치 우선"""
    # 1순위: 거래 키워드와 함께 등장하는 금액
    for m in AMOUNT_NEAR_KEYWORD.finditer(text):
        amount_str = m.group(2) or m.group(3)
        if not amount_str:
            continue
        if _is_blacklisted_context(text, m.start()):
            continue
        try:
            return int(amount_str.replace(",", ""))
        except ValueError:
            continue

    # 2순위: 단순 N원 매치 중 블랙리스트 컨텍스트 제외
    for m in AMOUNT_FALLBACK.finditer(text):
        if _is_blacklisted_context(text, m.start()):
            continue
        try:
            value = int(m.group(1).replace(",", ""))
            if value < 100:  # 100원 미만은 노이즈일 가능성 높음
                continue
            return value
        except ValueError:
            continue
    return None


def parse_transaction_type(text: str, source: str | None = None) -> str:
    if source in ("나이스정보통신", "토스페이먼츠", "헥토파이낸셜", "네이버페이"):
        if "취소" in text or "환불" in text:
            return "입금"
        return "출금"
    # 입금 키워드 우선 (환불/취소 포함)
    for kw in INCOME_KEYWORDS:
        if kw in text:
            return "입금"
    for kw in EXPENSE_KEYWORDS:
        if kw in text:
            return "출금"
    return "출금"


def parse_merchant(text: str, source: str) -> str:
    patterns = []
    if "나이스" in source:
        patterns += [
            r"([^\s]+(?:주식회사|㈜)?[^\s]+)(?:에서|에서의)\s*결제",
            r"님,\s*(.+?)에서\s*결제",
        ]
    if "토스" in source:
        patterns += [r"님,\s*(.+?)에서\s*결제한"]
    if "헥토" in source:
        patterns += [
            r"\(주\)([^\s]+)에서",
            r"님,\s*(.+?)에서",
        ]
    if "네이버" in source:
        patterns += [r"결제처[:\s]*([^\n\r]+)"]
    if "카카오" in source:
        patterns += [r"(?:가맹점|결제처|내역)[:\s]*([^\n\r]+)"]
    if "BC" in source:
        patterns += [r"(?:가맹점|사용처)[:\s]*([^\n\r]+)"]
    if "IBK" in source or "기업" in source:
        patterns += [r"(?:내용|적요)[:\s]*([^\n\r]+)"]
    if "현대" in source:
        patterns += [r"(?:가맹점|사용처|이용내역)[:\s]*([^\n\r]+)"]

    for pattern in patterns:
        m = re.search(pattern, text)
        if m:
            value = _clean_merchant_value(m.group(1))
            if value:
                return value[:50]
    return "알 수 없음"


def _clean_merchant_value(value: str) -> str:
    """가맹점 추출 결과에서 HTML 태그 잔재/엔티티/잡공백 제거"""
    if not value:
        return ""
    # HTML 태그 제거 (`</td>` 같은 잔재 정리)
    value = re.sub(r"<[^>]+>", " ", value)
    # HTML 엔티티 디코드
    value = html_module.unescape(value)
    # 제로폭/제어문자 제거
    value = re.sub(r"[​-‏﻿ ]", " ", value)
    # 공백 정리
    value = re.sub(r"\s+", " ", value).strip()
    return value


def guess_category(merchant: str, tx_type: str) -> str:
    if tx_type == "입금":
        return "수입"
    text = (merchant or "").lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                return category
    return "기타"


# ── 비거래 이메일 분류 ────────────────────────────────
def classify_non_transaction(sender: str, subject: str) -> str | None:
    """비거래 이메일을 카테고리 폴더로 분류. 매칭되는 폴더 이름 또는 None.

    제목 매칭은 공백 무시 — '배송 완료' / '배송완료' / '배 송 완 료' 모두 매칭.
    AliExpress 등 발신자가 제목 포맷팅을 다르게 쓰는 케이스 대응.
    """
    sender_lower = sender.lower()
    subject_lower = subject.lower()
    # 공백 제거 버전 (제목 매칭 정확도 향상)
    subject_nospace = re.sub(r"\s+", "", subject_lower)
    for category, folder, sender_kws, subject_kws in NON_TX_CATEGORIES:
        for kw in sender_kws:
            if kw.lower() in sender_lower:
                return folder
        for kw in subject_kws:
            kw_lower = kw.lower()
            # 원본 또는 공백 제거 버전에서 매칭
            if kw_lower in subject_lower or re.sub(r"\s+", "", kw_lower) in subject_nospace:
                return folder
    return None


# ── BC카드 월간 명세서 PDF 파싱 ────────────────────────
def is_statement_email(subject: str) -> bool:
    # 명세서 안내 / 수령방법 변경 안내성 메일은 제외
    if any(x in subject for x in ("수령방법", "신청완료", "이메일로 신청", "안내드립니다")):
        return False
    return (
        "이용대금명세서" in subject
        or "이용대금 명세서" in subject
        or "이메일명세서" in subject
        or "e-메일명세서" in subject
        or "이용대금" in subject
        or "명세서 재발송" in subject
        or "명세서가 도착" in subject
    )


def parse_kb_email_html(html_text: str) -> list[dict]:
    """KB국민카드 이메일 명세서 HTML에서 거래 목록 추출.

    KB는 본문 HTML 안의 `var list_pe01Json = [...]` JavaScript 변수에
    개별 거래(<tr>) 데이터를 임베드한다. PDF 첨부도 아니고 일반 본문
    텍스트도 아니라 별도 파서 필요.

    각 거래의 cell 순서: [날짜(YY.MM.DD), 매입처구분, 결제유형, 가맹점,
    공란, 금액, ...]
    """
    transactions = []

    # 요약 (pe00) — 청구월·결제일·이용기간
    bill_month = ""
    m_summary = re.search(r"var\s+list_pe00Json\s*=\s*\[(.*?)\];", html_text, re.DOTALL)
    if m_summary:
        mm = re.search(r'"결제년월일"\s*:\s*"([^"]*)"', m_summary.group(1))
        if mm:
            bill_month = mm.group(1).strip()

    # 개별 거래 (pe01)
    m_tx = re.search(r"var\s+list_pe01Json\s*=\s*\[(.*?)\];", html_text, re.DOTALL)
    if not m_tx:
        return transactions
    body = m_tx.group(1)
    # KB 포맷: key는 double quote("data"), HTML 값은 single quote로 감쌈
    rows = re.findall(r'"data"\s*:\s*\'(<tr>.*?</tr>)\'', body, re.DOTALL)
    if not rows:
        # 다른 견적 (KB 포맷 변경 대비)
        rows = re.findall(r"'data'\s*:\s*'(<tr>.*?</tr>)'", body, re.DOTALL)
        if not rows:
            rows = re.findall(r'"data"\s*:\s*"(<tr>.*?</tr>)"', body, re.DOTALL)

    for row in rows:
        tds = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL)
        cells = []
        for td in tds:
            txt = re.sub(r"<[^>]+>", "", td)
            txt = html_module.unescape(txt).strip()
            txt = re.sub(r"\s+", " ", txt)
            cells.append(txt)
        # KB 정상 거래 행은 td 13개. 8개 이하는 합계/안내 행 (skip)
        if len(cells) < 9 or not cells[0]:
            continue
        # 날짜 26.04.30 → 2026-04-30
        date_str = cells[0]
        parts = date_str.split(".")
        if len(parts) != 3:
            continue
        try:
            yy = int(parts[0])
            month = int(parts[1])
            day = int(parts[2])
        except ValueError:
            continue
        year = 2000 + yy if yy < 100 else yy
        date_iso = f"{year:04d}-{month:02d}-{day:02d}"

        # 회계 기준: td8(이번달 청구 분담분) 우선 사용.
        # 할부 거래는 td5=전체 이용금액 / td8=N개월 분담분으로 다름.
        # 일시불은 td5 == td8. 현대카드 파서와 같은 분담 모델 → IBK 자동이체와 정합.
        amt_str = re.sub(r"[^\d\-]", "", cells[8]) or re.sub(r"[^\d\-]", "", cells[5])
        try:
            amount = int(amt_str) if amt_str and amt_str != "-" else 0
        except ValueError:
            continue
        if amount <= 0:
            continue

        merchant = cells[3] or "KB카드 사용"
        pay_type = cells[2] or ""
        kind = cells[1] or ""
        origin_parts = [p for p in ["KB", pay_type, kind, f"청구 {bill_month}" if bill_month else ""] if p]
        transactions.append({
            "날짜": date_iso,
            "시간": "",
            "출처": "KB카드",
            "유형": "출금",
            "금액": amount,
            "내역": merchant[:50],
            "카테고리": guess_category(merchant, "출금"),
            "원문": " | ".join(origin_parts)[:100],
        })
    return transactions


# 은행/카드사가 발송하지만 거래 알림이 아닌 메일 (안내/공지/한도/약관 등)
_NON_TX_SUBJECT_KEYWORDS = (
    "한도초과", "한도 초과", "안내", "공지", "약관", "변경 안내",
    "이벤트", "혜택", "당첨", "이용 안내", "이용안내",
    "비밀번호", "보안", "위험자산", "투자 위험", "고지서",
    "프로모션", "추첨", "당첨자",
)


def is_non_transaction_subject(subject: str) -> bool:
    """은행/카드사 발신이지만 거래가 아닌 안내성 메일인지"""
    if not subject:
        return False
    return any(kw in subject for kw in _NON_TX_SUBJECT_KEYWORDS)


def get_pdf_attachment(msg) -> tuple:
    """첫 번째 PDF 첨부의 (파일명, 바이트) 반환"""
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename_raw = part.get_filename()
        filename = decode_str(filename_raw) if filename_raw else ""
        ctype = part.get_content_type()
        if ctype == "application/pdf" or filename.lower().endswith(".pdf"):
            try:
                payload = part.get_payload(decode=True)
            except Exception:
                payload = None
            if payload:
                return filename, payload
    return None, None


def get_xlsx_attachment(msg) -> tuple:
    """첫 번째 xlsx 첨부의 (파일명, 바이트) 반환"""
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        filename_raw = part.get_filename()
        filename = decode_str(filename_raw) if filename_raw else ""
        if filename.lower().endswith(".xlsx"):
            try:
                payload = part.get_payload(decode=True)
            except Exception:
                payload = None
            if payload:
                return filename, payload
    return None, None


def is_kakao_export_email(subject: str) -> bool:
    """카카오뱅크 '거래내역 엑셀' 내보내기 메일인지 (공백 변형 허용)."""
    s = re.sub(r"\s+", "", subject or "")
    return "거래내역엑셀" in s or ("거래내역" in s and "요청하신" in s)


def parse_kakao_export_xlsx(xlsx_bytes: bytes, password: str = "") -> list:
    """카카오뱅크 내보내기 xlsx → 표준 거래 리스트.

    첨부는 보통 생년월일 6자리로 암호화되어 있다 (KAKAO_XLSX_PASSWORD).
    출력 형식(내역·원문)은 app.py의 수동 업로드 파서와 동일하게 맞춰
    시트 중복 키(날짜_출처_금액_내역)가 양쪽 경로에서 일치하도록 한다.
    """
    import io

    buf = io.BytesIO(xlsx_bytes)
    try:
        import msoffcrypto
        of = msoffcrypto.OfficeFile(buf)
        if of.is_encrypted():
            if not password:
                raise ValueError("암호화 xlsx — KAKAO_XLSX_PASSWORD 시크릿 필요")
            decrypted = io.BytesIO()
            of.load_key(password=password)
            of.decrypt(decrypted)
            decrypted.seek(0)
            buf = decrypted
        else:
            buf.seek(0)
    except ImportError:
        buf.seek(0)

    from openpyxl import load_workbook
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    header_idx = None
    for i, row in enumerate(rows[:20]):
        cells = [_norm_header_cell(c) for c in row]
        if "거래일시" in cells and "거래금액" in cells:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("카카오뱅크 xlsx 헤더 행(거래일시+거래금액)을 찾지 못함")

    header = [_norm_header_cell(c) for c in rows[header_idx]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_dt = col("거래일시")
    i_kind = col("구분")
    i_amt = col("거래금액")
    i_bal = col("거래후잔액", "잔액")
    i_txkind = col("거래구분")
    i_content = col("내용")
    i_memo = col("메모")

    def _cell(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def _to_int(s):
        s = re.sub(r"[^\d\-]", "", str(s or ""))
        try:
            return int(s) if s and s != "-" else 0
        except ValueError:
            return 0

    txs = []
    for row in rows[header_idx + 1:]:
        raw_dt = row[i_dt] if i_dt is not None and i_dt < len(row) else None
        if raw_dt is None:
            continue
        if isinstance(raw_dt, datetime):
            dt = raw_dt
        else:
            dt = None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M:%S",
                        "%Y-%m-%d %H:%M", "%Y.%m.%d %H:%M", "%Y-%m-%d", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(str(raw_dt).strip(), fmt)
                    break
                except ValueError:
                    continue
            if dt is None:
                continue

        amt_signed = _to_int(_cell(row, i_amt))
        if amt_signed == 0:
            continue
        kind = _cell(row, i_kind)
        if kind == "출금":
            tx_type = "출금"
        elif kind == "입금":
            tx_type = "입금"
        else:
            tx_type = "출금" if amt_signed < 0 else "입금"

        content = _cell(row, i_content)
        txkind = _cell(row, i_txkind)
        memo = _cell(row, i_memo)
        balance = _to_int(_cell(row, i_bal)) if i_bal is not None else ""
        merchant = (content or txkind or "카카오뱅크 거래")[:50]
        origin = " | ".join(p for p in ["카카오뱅크", txkind, memo] if p)

        txs.append({
            "날짜": dt.strftime("%Y-%m-%d"),
            "시간": dt.strftime("%H:%M"),
            "출처": "카카오뱅크",
            "유형": tx_type,
            "금액": abs(amt_signed),
            "내역": merchant,
            "카테고리": guess_category(merchant, tx_type),
            "원문": origin[:100],
            "잔액": balance,
        })
    return txs


def process_kakao_exports(mail, folders: list, dest_folders: dict) -> tuple:
    """카카오뱅크 '거래내역 엑셀' 내보내기 메일 처리. (transactions, moved) 반환.

    사용자가 카뱅 앱에서 '내보내기 → 이메일'만 누르면 첨부 xlsx를 자동
    파싱해 시트에 넣는다 — 다운로드·업로드 없이 앱 버튼 하나로 수집 완료.
    암호 해독 실패(시크릿 미설정 등) 시 INBOX에 남겨 다음 실행에서 재시도.
    """
    transactions = []
    moved = 0
    since = (datetime.now() - timedelta(days=STATEMENT_LOOKBACK_DAYS)).strftime("%d-%b-%Y")

    for folder in folders:
        try:
            if mail.select(f'"{folder}"', readonly=False)[0] != "OK":
                continue
            _, data = mail.search(None, f'(SINCE "{since}" FROM "kakaobank.com")')
            eids = data[0].split() if data and data[0] else []
        except Exception as exc:
            print(f"카뱅 내보내기 검색 실패 ({folder}): {exc}")
            continue

        for eid in eids:
            try:
                _, msg_data = mail.fetch(eid, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                subject = decode_str(msg.get("Subject", ""))
            except Exception as exc:
                print(f"카뱅 내보내기 FETCH 실패 (eid={eid}): {exc}")
                continue
            if not is_kakao_export_email(subject):
                continue

            fname, xlsx = get_xlsx_attachment(msg)
            if not xlsx:
                print(f"  · 카뱅 내보내기 첨부 없음: {subject[:50]}")
                continue
            print(f"  · 카카오뱅크 내보내기 xlsx 파싱 중: ({len(xlsx):,}B)")
            try:
                txs = parse_kakao_export_xlsx(xlsx, KAKAO_XLSX_PASSWORD)
            except Exception as exc:
                print(f"    → 파싱 실패, INBOX 유지: {exc}")
                continue
            if txs:
                transactions.extend(txs)
                print(f"    → {len(txs)}건 추출")
            if ENABLE_EMAIL_CLEANUP and "처리완료" in dest_folders:
                if move_email(mail, eid, dest_folders["처리완료"]):
                    moved += 1

        if ENABLE_EMAIL_CLEANUP:
            try:
                mail.expunge()
            except Exception:
                pass
    return transactions, moved


def _norm_header_cell(s) -> str:
    return re.sub(r"\s+", "", str(s or ""))


def _find_col_idx(header: list, candidates: list):
    for i, h in enumerate(header):
        for c in candidates:
            if c in h:
                return i
    return None


def normalize_statement_date(raw: str, statement_year: int, statement_month: int):
    """다양한 날짜 포맷 → YYYY-MM-DD. 명세서 발행 월보다 큰 월은 전년도로 가정."""
    if not raw:
        return None
    raw = str(raw).strip()
    m = re.match(r"(\d{4})[\.\-/]\s*(\d{1,2})[\.\-/]\s*(\d{1,2})", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"(\d{1,2})[\.\-/]\s*(\d{1,2})", raw)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return None
        year = statement_year - 1 if month > statement_month else statement_year
        return f"{year}-{month:02d}-{day:02d}"
    return None


def normalize_statement_amount(raw: str):
    if raw is None:
        return None
    s = str(raw).replace(",", "").replace("원", "").replace(" ", "")
    if not s:
        return None
    sign = 1
    if s.startswith("-"):
        sign = -1
        s = s[1:]
    elif s.startswith("(") and s.endswith(")"):
        sign = -1
        s = s[1:-1]
    try:
        return int(s) * sign
    except ValueError:
        return None


_DATE_CELL_RE = re.compile(r"^\s*\d{1,2}[/.\-]\d{1,2}\s*$")

# 명세서에서 거래가 아닌 행을 식별하는 키워드 (가맹점명에 등장하면 skip)
_STATEMENT_SKIP_MERCHANT_KEYWORDS = ("소계", "합계", "(본인)", "(신용)", "(체크)")


def _detect_header_and_data_start(table: list):
    """다중 행 헤더 + 데이터 시작 위치 감지.
    헤더는 컬럼별로 세로 join하여 키워드 매칭에 사용한다.
    반환: (flat_header, data_start_idx) 또는 (None, None) — 데이터 행 없음."""
    if not table:
        return None, None
    data_start = None
    for i, row in enumerate(table):
        if not row:
            continue
        first = row[0] if len(row) > 0 else ""
        if _DATE_CELL_RE.match(str(first or "").strip()):
            data_start = i
            break
    if data_start is None or data_start == 0:
        return None, None

    header_rows = table[:data_start]
    max_cols = max((len(r or []) for r in header_rows), default=0)
    flat = []
    for col in range(max_cols):
        parts = []
        for r in header_rows:
            cell = r[col] if r and col < len(r) else None
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if cell_str:
                parts.append(cell_str)
        flat.append(_norm_header_cell(" ".join(parts)))
    return flat, data_start


def parse_statement_table(table: list, statement_year: int, statement_month: int) -> list:
    """pdfplumber.extract_tables()의 한 표를 거래 리스트로 변환.
    BC카드 명세서는 다중 행 헤더 + 그룹/소계 행이 섞여 있으므로 이를 모두 처리한다."""
    if not table or len(table) < 2:
        return []

    flat_header, data_start = _detect_header_and_data_start(table)
    if not flat_header or data_start is None:
        return []

    date_idx = _find_col_idx(flat_header, ["이용일자", "이용일", "사용일", "거래일", "승인일", "매출일"])
    merchant_idx = _find_col_idx(flat_header, [
        "가맹점(은행)명", "가맹점명", "가맹점", "이용처", "사용처", "이용내역", "내용"
    ])
    # 청구 기준 우선순위: 원금(KRW) > 청구금액/결제금액 > 이용금액
    amount_primary = _find_col_idx(flat_header, ["원금(KRW)", "원금"])
    amount_secondary = _find_col_idx(flat_header, ["청구금액", "결제금액", "승인금액"])
    amount_fallback = _find_col_idx(flat_header, ["이용금액", "이용 금액", "금액"])
    amount_candidates = [i for i in (amount_primary, amount_secondary, amount_fallback) if i is not None]
    if date_idx is None or not amount_candidates:
        return []

    out = []
    for row in table[data_start:]:
        if not row:
            continue
        merchant_raw = ""
        if merchant_idx is not None and merchant_idx < len(row):
            merchant_raw = str(row[merchant_idx] or "").strip()
        if not merchant_raw:
            continue
        if any(kw in merchant_raw for kw in _STATEMENT_SKIP_MERCHANT_KEYWORDS):
            continue

        date_raw = str(row[date_idx] or "").strip() if date_idx < len(row) else ""
        date = normalize_statement_date(date_raw, statement_year, statement_month)
        if not date:
            continue

        # 우선순위에 따라 첫 비어있지 않은 금액 셀 사용
        amount = None
        for idx in amount_candidates:
            if idx >= len(row):
                continue
            cell = row[idx]
            if cell is None or not str(cell).strip():
                continue
            amount = normalize_statement_amount(cell)
            if amount is not None:
                break
        if amount is None or amount == 0:
            continue

        merchant_clean = re.sub(r"\s+", " ", merchant_raw)[:50]
        tx_type = "입금" if amount < 0 else "출금"
        out.append({
            "날짜": date,
            "시간": "",
            "출처": "BC카드",
            "유형": tx_type,
            "금액": abs(amount),
            "내역": merchant_clean,
            "카테고리": guess_category(merchant_clean, tx_type),
            "원문": "BC카드 월간명세서",
        })
    return out


_AMOUNT_TOKEN_RE = re.compile(r"-?[\d]{1,3}(?:,[\d]{3})+(?:\.\d+)?")  # 12,345 / -1,234,567
_AMOUNT_TOKEN_BARE_RE = re.compile(r"-?[\d]{4,}(?:\.\d+)?")  # 4자리 이상 무콤마 정수


def _is_amount_token(tok: str) -> bool:
    """토큰이 거래 금액으로 보이는지 (콤마 포함, 4자리+ 정수, 또는 0/0원)"""
    if not tok:
        return False
    if tok in ("0", "0원", "(0)"):
        return True
    if tok.endswith("원"):
        tok = tok[:-1]
    if _AMOUNT_TOKEN_RE.fullmatch(tok):
        return True
    if _AMOUNT_TOKEN_BARE_RE.fullmatch(tok):
        return True
    if tok.startswith("(") and tok.endswith(")") and _AMOUNT_TOKEN_RE.fullmatch(tok[1:-1]):
        return True
    return False


_LINE_DATE_RE = re.compile(r"^\s*(\d{1,2})[\.\-/](\d{1,2})\s+(.+)$")


def parse_statement_text(text: str, statement_year: int, statement_month: int) -> list:
    """표 추출 실패 시 텍스트 라인 단위 fallback.
    BC카드 명세서 한 줄에는 여러 숫자(이용금액/할부개월/회차/원금/수수료/할인/잔액)가 있다.
    토큰화하여 가맹점(첫 amount 토큰 이전의 텍스트) + 첫 비-0 amount 토큰을 사용한다.
    할부개월/회차 같은 1~3자리 무콤마 정수는 금액에서 제외한다."""
    if not text:
        return []
    out = []
    for line in text.splitlines():
        m = _LINE_DATE_RE.match(line)
        if not m:
            continue
        month, day = int(m.group(1)), int(m.group(2))
        if not (1 <= month <= 12 and 1 <= day <= 31):
            continue

        tokens = m.group(3).split()
        if not tokens:
            continue

        merchant_parts = []
        amounts = []
        for tok in tokens:
            if _is_amount_token(tok):
                amounts.append(tok)
            elif amounts:
                # 첫 amount 이후의 비-amount 토큰은 metadata (특별서비스 "면제" 등) → 무시
                continue
            else:
                merchant_parts.append(tok)

        merchant = " ".join(merchant_parts).strip()
        if not merchant:
            continue
        if any(kw in merchant for kw in _STATEMENT_SKIP_MERCHANT_KEYWORDS):
            continue

        chosen = None
        for tok in amounts:
            v = normalize_statement_amount(tok)
            if v is None or v == 0:
                continue
            chosen = v
            break
        if chosen is None:
            continue

        year = statement_year - 1 if month > statement_month else statement_year
        date = f"{year}-{month:02d}-{day:02d}"
        merchant_clean = re.sub(r"\s+", " ", merchant)[:50]
        tx_type = "입금" if chosen < 0 else "출금"
        out.append({
            "날짜": date,
            "시간": "",
            "출처": "BC카드",
            "유형": tx_type,
            "금액": abs(chosen),
            "내역": merchant_clean,
            "카테고리": guess_category(merchant_clean, tx_type),
            "원문": "BC카드 월간명세서",
        })
    return out


def _extract_pdf_text_pymupdf(pdf_bytes: bytes, password: str):
    """PyMuPDF로 페이지별 텍스트 추출. 한국어 PDF의 ToUnicode 매핑 문제에 견고.
    반환: list[str] (페이지별), 또는 실패 시 None"""
    try:
        import fitz  # pymupdf
    except ImportError:
        return None
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        print(f"    [pymupdf] 열기 실패: {exc}")
        return None
    try:
        if doc.is_encrypted:
            ok = doc.authenticate(password)
            if not ok:
                print("    [pymupdf] 비밀번호 인증 실패")
                return None
        pages = []
        for page in doc:
            try:
                pages.append(page.get_text() or "")
            except Exception:
                pages.append("")
        return pages
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _extract_pdf_text_pymupdf_ocr(pdf_bytes: bytes, password: str):
    """PyMuPDF + Tesseract OCR로 페이지 텍스트 추출. 시스템에 tesseract + 한국어 데이터 필요."""
    try:
        import fitz
    except ImportError:
        print("    [ocr] pymupdf 미설치")
        return None
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        print(f"    [ocr] PDF 열기 실패: {exc}")
        return None
    try:
        if doc.is_encrypted and not doc.authenticate(password):
            print("    [ocr] 비밀번호 인증 실패")
            return None
        pages = []
        for i, page in enumerate(doc):
            try:
                tp = page.get_textpage_ocr(language="kor+eng", dpi=200, full=True)
                pages.append(page.get_text(textpage=tp) or "")
            except Exception as exc:
                print(f"    [ocr] page {i + 1} 실패: {exc}")
                pages.append("")
        return pages
    finally:
        try:
            doc.close()
        except Exception:
            pass


_HYBRID_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}$")
# 금액 토큰: 콤마 포함, 또는 3자리 이상 정수 (할부개월·회차 1~2자리는 X 범위로 걸러짐)
_HYBRID_AMOUNT_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})+$|^-?\d{3,}$")
_HANGUL_BETWEEN_SPACES = re.compile(r"([가-힣])\s+([가-힣])")
_OCR_NOISE = {"이", "|", ":", ";", "{", "}", "-", "ｌ", "！", "·", "‧"}


def _squeeze_korean_spaces(s: str) -> str:
    """OCR 결과의 한글 글자 사이 단일 공백 합치기 (`굿 모 닝` → `굿모닝`)"""
    if not s:
        return s
    prev = None
    while prev != s:
        prev = s
        s = _HANGUL_BETWEEN_SPACES.sub(r"\1\2", s)
    s = re.sub(r"\s+", " ", s).strip()
    # 괄호 안 공백 정리: ( 주 ) → (주)
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    return s


def _ocr_page_words(page, dpi: int, scale: float) -> list:
    """PDF 페이지를 PNG로 렌더링 후 Tesseract TSV로 OCR.
    반환: [(x_pt, y_pt_center, text), ...] — PDF 좌표계로 변환된 단어 리스트."""
    import os
    import subprocess
    import tempfile

    try:
        pix = page.get_pixmap(dpi=dpi)
    except Exception as exc:
        print(f"    [hybrid] 렌더링 실패: {exc}")
        return []

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        pix.save(f.name)
        png_path = f.name

    try:
        result = subprocess.run(
            ["tesseract", png_path, "-", "-l", "kor+eng", "--psm", "4", "tsv"],
            capture_output=True, text=True, timeout=120,
        )
        tsv = result.stdout
    except FileNotFoundError:
        print("    [hybrid] tesseract 바이너리 없음")
        return []
    except subprocess.TimeoutExpired:
        print("    [hybrid] tesseract 시간 초과")
        return []
    finally:
        try:
            os.unlink(png_path)
        except OSError:
            pass

    words = []
    for line in tsv.splitlines()[1:]:
        parts = line.split("\t")
        if len(parts) < 12:
            continue
        try:
            left = int(parts[6])
            top = int(parts[7])
            height = int(parts[9])
            conf = float(parts[10])
            text = parts[11].strip()
        except (ValueError, IndexError):
            continue
        if not text or text in _OCR_NOISE or conf < 30:
            continue
        # 순수 숫자 토큰은 OCR 부정확 — PDF text로 가져옴
        if re.fullmatch(r"[\d,.\-]+", text):
            continue
        words.append((left / scale, (top + height / 2) / scale, text))
    return words


def _extract_pdf_hybrid(pdf_bytes: bytes, password: str,
                       statement_year: int, statement_month: int) -> list:
    """PDF 텍스트(날짜·금액) + OCR(가맹점명) 하이브리드 추출.
    한국어 ToUnicode가 일부 폰트만 깨진 BC카드 명세서에 최적.
    PDF text에서 같은 라인(같은 Y) 안의 [date_x_end ~ first_amount_x] X 범위에 있는
    OCR 단어들을 모아 가맹점명으로 사용."""
    try:
        import fitz
    except ImportError:
        return []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        print(f"    [hybrid] PDF 열기 실패: {exc}")
        return []
    try:
        if doc.is_encrypted and not doc.authenticate(password):
            print("    [hybrid] 비밀번호 인증 실패")
            return []

        DPI = 300
        SCALE = DPI / 72.0
        all_txs = []

        for page_idx in range(doc.page_count):
            page = doc[page_idx]
            try:
                words = page.get_text("words") or []
            except Exception:
                words = []
            if not words:
                continue

            # 같은 Y 좌표에 있는 단어들을 묶어 PDF 행 추출
            by_y = {}
            for x0, y0, x1, y1, txt, *_ in words:
                key = round(y0)
                by_y.setdefault(key, []).append((x0, x1, y0, y1, txt))

            # BC카드 명세서 컬럼 X 좌표 (300dpi 기준 PDF pt):
            #  - X≈260 : "이용금액" = 체크(직불) 결제
            #  - X≈351 : "원금(KRW)" = 신용 결제 (일시불 또는 할부 이번달 분담)
            # 같은 행에 둘 다 있으면 → 신용 할부 (X=260은 원본 거래액 정보, X=351이 이번달 청구)
            X_CHECK_RANGE = (250, 290)
            X_CREDIT_RANGE = (340, 380)

            pdf_rows = []
            for y_key in sorted(by_y):
                ws = sorted(by_y[y_key], key=lambda w: w[0])
                toks = [w[4] for w in ws]
                if not toks or not _HYBRID_DATE_RE.match(toks[0]):
                    continue
                check_amt = None
                credit_amt = None
                for w in ws[1:]:
                    x, _, _, _, txt = w
                    if not _HYBRID_AMOUNT_RE.match(txt):
                        continue
                    if X_CHECK_RANGE[0] <= x <= X_CHECK_RANGE[1] and check_amt is None:
                        check_amt = (x, txt)
                    elif X_CREDIT_RANGE[0] <= x <= X_CREDIT_RANGE[1] and credit_amt is None:
                        credit_amt = (x, txt)
                # 신용(원금) 컬럼이 있으면 우선 사용 (할부 분담액 또는 신용 일시불)
                if credit_amt:
                    use_x, use_amt = credit_amt
                    card_kind = "신용"
                elif check_amt:
                    use_x, use_amt = check_amt
                    card_kind = "체크"
                else:
                    continue
                # 가맹점 텍스트가 끝나는 X = 첫 금액의 X (체크/신용 중 X 작은 쪽)
                first_x = min(filter(None, [
                    check_amt[0] if check_amt else None,
                    credit_amt[0] if credit_amt else None,
                ]))
                pdf_rows.append({
                    "y_center": (ws[0][2] + ws[0][3]) / 2,
                    "date_str": toks[0],
                    "date_x_end": ws[0][1],
                    "first_amount_x": first_x,
                    "amount_str": use_amt,
                    "card_kind": card_kind,
                })

            if not pdf_rows:
                print(f"    [hybrid] page {page_idx + 1}: PDF 거래 행 0개, OCR 생략")
                continue

            ocr_words = _ocr_page_words(page, DPI, SCALE)
            if not ocr_words:
                print(f"    [hybrid] page {page_idx + 1}: OCR 0건")
                continue

            page_txs = []
            for r in pdf_rows:
                cands = [
                    (x, t) for x, y, t in ocr_words
                    if abs(y - r["y_center"]) <= 7
                    and r["date_x_end"] + 2 < x < r["first_amount_x"] - 2
                ]
                cands.sort(key=lambda c: c[0])
                merchant = _squeeze_korean_spaces(" ".join(t for _, t in cands))[:50]

                m = re.match(r"(\d{1,2})/(\d{1,2})", r["date_str"])
                if not m:
                    continue
                month, day = int(m.group(1)), int(m.group(2))
                if not (1 <= month <= 12 and 1 <= day <= 31):
                    continue
                year = statement_year - 1 if month > statement_month else statement_year
                date_iso = f"{year}-{month:02d}-{day:02d}"

                amount = normalize_statement_amount(r["amount_str"])
                if amount is None or amount == 0:
                    continue

                if not merchant:
                    merchant = "알 수 없음"
                tx_type = "입금" if amount < 0 else "출금"
                source = f"BC카드({r['card_kind']})"
                page_txs.append({
                    "날짜": date_iso,
                    "시간": "",
                    "출처": source,
                    "유형": tx_type,
                    "금액": abs(amount),
                    "내역": merchant,
                    "카테고리": guess_category(merchant, tx_type),
                    "원문": "BC카드 월간명세서",
                })
            all_txs.extend(page_txs)
            print(f"    [hybrid] page {page_idx + 1}: txs={len(page_txs)}")
        return all_txs
    finally:
        try:
            doc.close()
        except Exception:
            pass


# 가맹점명이 깨진 인코딩인지 판정용 — Latin-1 supplement, 기호류 영역.
# Korean PDF의 ToUnicode CMap 깨짐 시 흔히 나타나는 mojibake 문자 영역.
_GARBLED_RANGES = (
    (0x00A0, 0x00FF),  # Latin-1 supplement (¡«¬©Æ°÷ 등)
    (0x0100, 0x017F),  # Latin Extended-A (Ÿı 등)
    (0x0250, 0x02AF),  # IPA Extensions
    (0x02B0, 0x02FF),  # Spacing Modifier Letters (ˆ˝˜ 등)
    (0x2030, 0x205F),  # General Punctuation (‰⁄… 등, 0x2018-201F 스마트인용부호 제외)
    (0x2100, 0x214F),  # Letterlike Symbols
    (0x2150, 0x218F),  # Number Forms
    (0x2200, 0x22FF),  # Mathematical Operators (∂∞∫≠ 등)
    (0x2300, 0x23FF),  # Miscellaneous Technical
    (0x25A0, 0x25FF),  # Geometric Shapes (◊ 등)
    (0x2500, 0x257F),  # Box Drawing
)


def _looks_garbled(merchants: list) -> bool:
    """가맹점명 컬렉션이 한국어가 아닌 의심 문자(Latin-1 등)로 채워져 있는지.
    한 가맹점에 의심 문자가 1개라도 있으면 'suspect', 전체의 1/3 이상이 suspect면 깨짐."""
    if not merchants:
        return False

    def has_suspect(s: str) -> bool:
        for ch in s:
            cp = ord(ch)
            for lo, hi in _GARBLED_RANGES:
                if lo <= cp <= hi:
                    return True
        return False

    suspect_count = sum(1 for m in merchants if has_suspect(m or ""))
    return suspect_count >= max(2, len(merchants) // 3)


def _extract_pdf_with_pdfplumber(pdf_bytes: bytes, password: str,
                                  statement_year: int, statement_month: int) -> list:
    """pdfplumber 기반 추출 (표 우선, 텍스트 fallback). PyMuPDF 실패 시 사용."""
    import io
    try:
        import pdfplumber
    except ImportError:
        print("    [pdfplumber] 미설치")
        return []
    try:
        pdf = pdfplumber.open(io.BytesIO(pdf_bytes), password=password)
    except Exception as exc:
        print(f"    [pdfplumber] 열기 실패: {exc}")
        return []
    all_txs = []
    try:
        for page_idx, page in enumerate(pdf.pages):
            tables = []
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            page_table_count = 0
            for table in tables:
                page_txs = parse_statement_table(table, statement_year, statement_month)
                if page_txs:
                    page_table_count += len(page_txs)
                    all_txs.extend(page_txs)
            page_text_count = 0
            if page_table_count == 0:
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                fallback_txs = parse_statement_text(text, statement_year, statement_month)
                page_text_count = len(fallback_txs)
                all_txs.extend(fallback_txs)
            print(
                f"    [pdfplumber] page {page_idx + 1}: tables={len(tables)}, "
                f"table_txs={page_table_count}, fallback_txs={page_text_count}"
            )
    finally:
        try:
            pdf.close()
        except Exception:
            pass
    return all_txs


def parse_pdf_transactions(pdf_bytes: bytes, password: str,
                           statement_year: int, statement_month: int) -> list:
    """비밀번호 PDF에서 거래 추출. 4-tier 전략:
    0) Hybrid (PDF text 날짜/금액 + Tesseract OCR 가맹점 Y좌표 매칭) — BC카드 최적
    1) PyMuPDF 일반 텍스트 (정상 PDF)
    2) pdfplumber (표 + 텍스트 fallback)
    3) PyMuPDF + Tesseract OCR (전체 OCR fallback)
    각 결과의 가맹점명이 깨진 인코딩으로 판정되면 자동으로 다음 티어로 넘어감."""

    # 0순위: Hybrid (BC카드 명세서처럼 가맹점만 깨진 PDF용)
    hybrid_txs = _extract_pdf_hybrid(pdf_bytes, password, statement_year, statement_month)
    if hybrid_txs and not _looks_garbled([t["내역"] for t in hybrid_txs]):
        return _dedup_pdf_transactions(hybrid_txs)
    if hybrid_txs:
        print("    [hybrid] 추출됐으나 인코딩 깨짐 감지")

    # 1순위: PyMuPDF 일반 텍스트
    pages = _extract_pdf_text_pymupdf(pdf_bytes, password)
    pymupdf_txs = []
    if pages is not None:
        for i, text in enumerate(pages):
            page_txs = parse_statement_text(text, statement_year, statement_month)
            print(f"    [pymupdf] page {i + 1}: txs={len(page_txs)}")
            pymupdf_txs.extend(page_txs)
    if pymupdf_txs and not _looks_garbled([t["내역"] for t in pymupdf_txs]):
        return _dedup_pdf_transactions(pymupdf_txs)
    if pymupdf_txs:
        print("    [pymupdf] 추출됐으나 인코딩 깨짐 감지")

    # 2순위: pdfplumber
    print("    pdfplumber 재시도")
    plumber_txs = _extract_pdf_with_pdfplumber(pdf_bytes, password, statement_year, statement_month)
    if plumber_txs and not _looks_garbled([t["내역"] for t in plumber_txs]):
        return _dedup_pdf_transactions(plumber_txs)
    if plumber_txs:
        print("    [pdfplumber] 추출됐으나 인코딩 깨짐 감지")

    # 3순위: PyMuPDF + Tesseract OCR
    print("    텍스트 레이어 추출 불가 → OCR fallback 시도")
    ocr_pages = _extract_pdf_text_pymupdf_ocr(pdf_bytes, password)
    if ocr_pages is None:
        print("    [ocr] 사용 불가, 깨진 결과라도 반환")
        # 인코딩 깨졌어도 pymupdf보다는 pdfplumber 결과가 정보량 많음
        return _dedup_pdf_transactions(plumber_txs or pymupdf_txs)
    ocr_txs = []
    for i, text in enumerate(ocr_pages):
        page_txs = parse_statement_text(text, statement_year, statement_month)
        print(f"    [ocr] page {i + 1}: txs={len(page_txs)}")
        ocr_txs.extend(page_txs)
    if ocr_txs:
        return _dedup_pdf_transactions(ocr_txs)
    print("    [ocr] 매칭 0건. PDF 포맷 진단 필요")
    return _dedup_pdf_transactions(plumber_txs or pymupdf_txs)


def _dedup_pdf_transactions(txs: list) -> list:
    seen = set()
    out = []
    for tx in txs:
        key = (tx["날짜"], tx["내역"], tx["금액"])
        if key in seen:
            continue
        seen.add(key)
        out.append(tx)
    return out


# ── IMAP 작업 ─────────────────────────────────────────
def connect_imap():
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    mail.login(NAVER_EMAIL, NAVER_APP_PW)
    return mail


def ensure_folder(mail, folder_name: str) -> str:
    """폴더가 없으면 생성하고 IMAP UTF-7 인코딩된 이름 반환"""
    encoded = imap_utf7_encode(folder_name)
    try:
        mail.create(f'"{encoded}"')
    except Exception:
        pass
    return encoded


def move_email(mail, eid, dest_encoded: str, mark_seen: bool = True) -> bool:
    """이메일을 dest_encoded 폴더로 복사 후 원본 삭제 표시. EXPUNGE는 호출자 책임"""
    try:
        if mark_seen:
            mail.store(eid, "+FLAGS", "\\Seen")
        result, _ = mail.copy(eid, f'"{dest_encoded}"')
        if result != "OK":
            return False
        mail.store(eid, "+FLAGS", "\\Deleted")
        return True
    except Exception as exc:
        print(f"이동 실패 (eid={eid}, dest={dest_encoded}): {exc}")
        return False


def fetch_recent_ids(mail, folder: str, hours: int) -> list:
    """폴더에서 최근 N시간 이내 이메일 ID 리스트"""
    since = (datetime.now() - timedelta(hours=hours)).strftime("%d-%b-%Y")
    status, _ = mail.select(f'"{folder}"', readonly=False)
    if status != "OK":
        return []
    _, data = mail.search(None, f'(SINCE "{since}")')
    return data[0].split() if data and data[0] else []


def process_folder(mail, folder: str, hours: int, dest_folders: dict,
                   cleanup_log: dict | None = None) -> tuple:
    """한 폴더를 처리: 거래 추출 + 비거래 분류 이동.
    cleanup_log: {카테고리: [제목,...]} — 정리 요약용. 전달 시 분류된 제목 수집.
    반환: (transactions, moved_counts dict)"""
    transactions = []
    moved = {"거래": 0, "쇼핑": 0, "SNS": 0, "뉴스레터": 0, "광고": 0}

    eids = fetch_recent_ids(mail, folder, hours)
    if not eids:
        return transactions, moved

    for eid in eids:
        try:
            _, msg_data = mail.fetch(eid, "(RFC822)")
            msg = email.message_from_bytes(msg_data[0][1])
            sender = decode_str(msg.get("From", ""))
            subject = decode_str(msg.get("Subject", ""))
            date_str = msg.get("Date", "")
        except Exception as exc:
            print(f"FETCH 실패 (folder={folder}, eid={eid}): {exc}")
            continue

        # 거래 이메일 매칭
        source = None
        for name, patterns in SENDER_PATTERNS.items():
            if any(p.lower() in sender.lower() for p in patterns):
                source = name
                break

        # BC카드 / KB카드 월간 명세서는 본문에 금액이 없거나 별도 HTML 구조라
        # 별도 패스에서 처리한다.
        if source == "BC카드" and is_statement_email(subject):
            continue
        if source == "KB카드" and is_statement_email(subject):
            continue

        # 은행/카드사 발신이지만 거래가 아닌 안내성 메일은 사전 차단
        if source and is_non_transaction_subject(subject):
            continue

        if source:
            try:
                body = get_email_body(msg)
                full_text = subject + "\n" + body

                amount = parse_amount(full_text)
                if amount is None:
                    print(f"  · 금액 파싱 실패: [{source}] {subject[:50]}")
                    continue
                if amount == 0:
                    # 0원 거래는 의미 없음 (면제/안내성)
                    continue

                tx_type = parse_transaction_type(full_text, source)
                merchant = parse_merchant(full_text, source)
                category = guess_category(merchant, tx_type)

                try:
                    dt = parsedate_to_datetime(date_str)
                    tx_date = dt.strftime("%Y-%m-%d")
                    tx_time = dt.strftime("%H:%M")
                except Exception:
                    tx_date = datetime.now().strftime("%Y-%m-%d")
                    tx_time = ""

                transactions.append({
                    "날짜": tx_date,
                    "시간": tx_time,
                    "출처": source,
                    "유형": tx_type,
                    "금액": amount,
                    "내역": merchant,
                    "카테고리": category,
                    "원문": subject[:100],
                })

                if ENABLE_EMAIL_CLEANUP and "처리완료" in dest_folders:
                    if move_email(mail, eid, dest_folders["처리완료"]):
                        moved["거래"] += 1
            except Exception as exc:
                print(f"거래 파싱 실패 (eid={eid}): {exc}")
            continue

        # 비거래 이메일 분류
        if not ENABLE_EMAIL_CLEANUP:
            continue
        category_folder = classify_non_transaction(sender, subject)
        if not category_folder:
            continue
        encoded = dest_folders.get(category_folder)
        if not encoded:
            continue
        # move_email은 mark_seen=True라 이동과 동시에 읽음 처리됨.
        # 홍보(광고) 메일도 여기서 읽음+이동으로 함께 정리된다.
        if move_email(mail, eid, encoded):
            # 카운트 키 매핑 + 요약 로그 수집
            for cat_name, folder_name, _, _ in NON_TX_CATEGORIES:
                if folder_name == category_folder:
                    moved[cat_name] = moved.get(cat_name, 0) + 1
                    if cleanup_log is not None:
                        cleanup_log.setdefault(cat_name, []).append(
                            decode_str(subject).strip()[:60]
                        )
                    break

    # EXPUNGE는 select가 풀린 후엔 무효 — 폴더별로 마지막에 호출
    if ENABLE_EMAIL_CLEANUP:
        try:
            mail.expunge()
        except Exception as exc:
            print(f"EXPUNGE 실패 ({folder}): {exc}")

    return transactions, moved


def prepare_dest_folders(mail) -> dict:
    """이동 대상 폴더들을 미리 생성하고 인코딩 매핑 반환"""
    if not ENABLE_EMAIL_CLEANUP:
        return {}
    mapping = {
        "처리완료": ensure_folder(mail, PROCESSED_FOLDER),
        AD_FOLDER: ensure_folder(mail, AD_FOLDER),
        SHOPPING_FOLDER: ensure_folder(mail, SHOPPING_FOLDER),
        NEWSLETTER_FOLDER: ensure_folder(mail, NEWSLETTER_FOLDER),
        SNS_FOLDER: ensure_folder(mail, SNS_FOLDER),
    }
    return mapping


def process_kb_statements(mail, folders: list, dest_folders: dict) -> tuple:
    """KB국민카드 이메일 명세서 HTML 처리. (transactions, moved_count) 반환.

    KB 명세서는 본문 HTML에 거래가 임베드되거나, '명세서 재발송' 메일의
    경우 본문은 안내문일 뿐이고 거래는 첨부 HTML(.html)에 들어있다.
    두 곳을 순서대로 시도한다.
    """
    transactions = []
    moved_count = 0

    since = (datetime.now() - timedelta(days=STATEMENT_LOOKBACK_DAYS)).strftime("%d-%b-%Y")

    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=False)
            if status != "OK":
                continue
            _, data = mail.search(None, f'(SINCE "{since}" FROM "kbcard")')
            eids = data[0].split() if data and data[0] else []
        except Exception as exc:
            print(f"KB 명세서 검색 실패 ({folder}): {exc}")
            continue

        for eid in eids:
            try:
                _, msg_data = mail.fetch(eid, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                subject = decode_str(msg.get("Subject", ""))
            except Exception as exc:
                print(f"KB 명세서 FETCH 실패 (eid={eid}): {exc}")
                continue

            if not is_statement_email(subject):
                continue

            # 본문 HTML과 .html 첨부 둘 다 수집
            html_candidates = []  # [(label, html_text)]
            for part in msg.walk():
                ctype = part.get_content_type()
                disp = header_str(part.get("Content-Disposition")).lower()
                fn = decode_str(part.get_filename() or "")
                is_attachment = "attachment" in disp or bool(fn)
                # 본문 text/html
                if ctype == "text/html" and not is_attachment:
                    try:
                        payload = part.get_payload(decode=True) or b""
                        cs = part.get_content_charset() or "cp949"
                        html_candidates.append(("body", payload.decode(cs, errors="replace")))
                    except Exception:
                        pass
                # 첨부 .html (KB 재발송 메일은 거래가 여기에)
                elif is_attachment and (ctype == "text/html" or fn.lower().endswith(".html") or fn.lower().endswith(".htm")):
                    try:
                        payload = part.get_payload(decode=True) or b""
                        # 첨부는 charset 헤더가 없으므로 cp949 우선 시도
                        for enc in ("cp949", "utf-8", "euc-kr"):
                            try:
                                html_candidates.append((f"attach:{fn or '?'}", payload.decode(enc)))
                                break
                            except UnicodeDecodeError:
                                continue
                    except Exception:
                        pass

            if not html_candidates:
                print(f"  · KB 명세서 HTML 없음 (본문·첨부): {subject[:60]}")
                continue

            print(f"  · KB 명세서 후보 {len(html_candidates)}개: {subject[:60]}")
            kb_txs = []
            chosen_label = None
            for label, html_text in html_candidates:
                txs = parse_kb_email_html(html_text)
                if txs:
                    kb_txs = txs
                    chosen_label = label
                    break

            if kb_txs:
                transactions.extend(kb_txs)
                print(f"    → {len(kb_txs)}건 추출 (소스: {chosen_label})")
                if ENABLE_EMAIL_CLEANUP and "처리완료" in dest_folders:
                    if move_email(mail, eid, dest_folders["처리완료"]):
                        moved_count += 1
            else:
                print("    → 추출 0건. KB HTML 포맷 변경 가능성, 로그 확인 필요")

        if ENABLE_EMAIL_CLEANUP:
            try:
                mail.expunge()
            except Exception as exc:
                print(f"KB 명세서 EXPUNGE 실패 ({folder}): {exc}")

    return transactions, moved_count


def process_statements(mail, folders: list, dest_folders: dict) -> tuple:
    """BC카드 / IBK BC카드 월간 명세서 PDF 처리. (transactions, moved_count) 반환.

    BC카드 본사 발신: bccard.com (보통 PDF 비밀번호 있음 → BC_PDF_PASSWORD 필요)
    IBK BC카드 발신: ibk.co.kr (보통 비암호 PDF)
    두 발신자 모두 동일한 BC카드 명세서 PDF 형식 사용.
    """
    transactions = []
    moved_count = 0

    since = (datetime.now() - timedelta(days=STATEMENT_LOOKBACK_DAYS)).strftime("%d-%b-%Y")
    # 두 발신자 도메인 모두 검색. IBK는 일반 거래 알림도 있으므로
    # is_statement_email + PDF 첨부 유무로 명세서만 추려낸다.
    sender_filters = ["bccard.com", "ibk.co.kr"]
    seen_eids = set()  # 폴더×eid 중복 처리 방지

    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=False)
            if status != "OK":
                continue
        except Exception as exc:
            print(f"명세서 폴더 선택 실패 ({folder}): {exc}")
            continue

        for sf in sender_filters:
            try:
                _, data = mail.search(None, f'(SINCE "{since}" FROM "{sf}")')
                eids = data[0].split() if data and data[0] else []
            except Exception as exc:
                print(f"명세서 검색 실패 ({folder}/{sf}): {exc}")
                continue

            for eid in eids:
                key = (folder, eid)
                if key in seen_eids:
                    continue
                seen_eids.add(key)
                try:
                    _, msg_data = mail.fetch(eid, "(RFC822)")
                    msg = email.message_from_bytes(msg_data[0][1])
                    subject = decode_str(msg.get("Subject", ""))
                    sender = decode_str(msg.get("From", ""))
                    date_str = msg.get("Date", "")
                except Exception as exc:
                    print(f"명세서 FETCH 실패 (eid={eid}): {exc}")
                    continue

                if not is_statement_email(subject):
                    continue

                try:
                    dt = parsedate_to_datetime(date_str)
                    s_year, s_month = dt.year, dt.month
                except Exception:
                    now = datetime.now()
                    s_year, s_month = now.year, now.month

                fname, pdf_bytes = get_pdf_attachment(msg)
                if not pdf_bytes:
                    print(f"  · 명세서 첨부 없음: {subject[:60]}")
                    continue

                # IBK BC카드는 비암호 PDF, BC카드 본사는 비번 필요.
                # 일단 비번 없이 시도, 실패하면 BC_PDF_PASSWORD로 재시도.
                # bccard.com / ibk.co.kr 둘 다 같은 BC카드 명세서 (사용자 카드는
                # IBK 발급 BC카드 하나) → 출처·파서·출력 동일하게 통일.
                print(
                    f"  · BC카드 명세서 PDF 파싱 중: {fname or '(이름 없음)'} "
                    f"({len(pdf_bytes):,}B)"
                )
                pdf_txs = parse_pdf_transactions(pdf_bytes, "", s_year, s_month)
                if not pdf_txs and BC_PDF_PASSWORD:
                    print("    [재시도] BC_PDF_PASSWORD로 복호화")
                    pdf_txs = parse_pdf_transactions(pdf_bytes, BC_PDF_PASSWORD, s_year, s_month)

                if pdf_txs:
                    transactions.extend(pdf_txs)
                    print(f"    → {len(pdf_txs)}건 추출 ({s_year}-{s_month:02d} 명세)")
                    if ENABLE_EMAIL_CLEANUP and "처리완료" in dest_folders:
                        if move_email(mail, eid, dest_folders["처리완료"]):
                            moved_count += 1
                else:
                    print("    → 추출 0건. PDF 포맷/비번 확인 필요")

        if ENABLE_EMAIL_CLEANUP:
            try:
                mail.expunge()
            except Exception as exc:
                print(f"명세서 EXPUNGE 실패 ({folder}): {exc}")

    return transactions, moved_count


# ── 주택금융공사 보금자리론 안내 메일 처리 ───────────
def parse_hf_loan_notice(text: str, base_year: int | None = None) -> dict | None:
    """주택금융공사 보금자리론 안내(SMS/이메일)에서 회차·원리금·잔액 추출.

    예: '06월 18일은 ... 116회차 ... 대출잔액 114,051,749원
         ... 총:729,960원(원금:465,517 이자:264,443) ...'
    app.py의 parse_loan_notice와 동일한 정규식 — email cron에서
    독립적으로 동작하기 위해 같은 로직을 복제한다.
    """
    if not text:
        return None
    out = {}
    m_date = re.search(r"(\d{1,2})월\s*(\d{1,2})일", text)
    if m_date:
        year = base_year or datetime.now().year
        out["납입일"] = f"{year:04d}-{int(m_date.group(1)):02d}-{int(m_date.group(2)):02d}"
    m_n = re.search(r"(\d+)\s*회차", text)
    if m_n:
        out["회차"] = int(m_n.group(1))
    m_bal = re.search(r"대출잔액\s*([\d,]+)\s*원", text)
    if m_bal:
        out["잔액"] = int(m_bal.group(1).replace(",", ""))
    m_pay = re.search(
        r"총\s*:?\s*([\d,]+)\s*원\s*\(\s*원금\s*:?\s*([\d,]+)\s*이자\s*:?\s*([\d,]+)",
        text,
    )
    if m_pay:
        out["납입액"] = int(m_pay.group(1).replace(",", ""))
        out["원금"] = int(m_pay.group(2).replace(",", ""))
        out["이자"] = int(m_pay.group(3).replace(",", ""))
    required = ("납입일", "회차", "납입액", "원금", "이자", "잔액")
    return out if all(k in out for k in required) else None


def process_hf_loan_emails(mail, folders: list, dest_folders: dict) -> int:
    """주택금융공사 안내 메일을 검색해 회차 정보를 '보금자리론' 워크시트에
    자동 추가. 반환: 추가된 회차 수.

    안내 본문(text 또는 HTML strip)에서 parse_hf_loan_notice로 추출.
    중복 회차는 skip. ENABLE_EMAIL_CLEANUP=true면 처리완료 폴더로 이동.
    """
    if not (GOOGLE_SHEET_ID and GOOGLE_CREDS_JSON):
        return 0
    since = (datetime.now() - timedelta(days=STATEMENT_LOOKBACK_DAYS)).strftime("%d-%b-%Y")
    added = 0
    # 보금자리론 워크시트 fetch는 첫 안내 발견 시 lazy하게
    loan_ws = None
    existing_rounds = set()

    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"', readonly=False)
            if status != "OK":
                continue
            _, data = mail.search(None, f'(SINCE "{since}" FROM "hf.go.kr")')
            eids = data[0].split() if data and data[0] else []
        except Exception as exc:
            print(f"주택금융공사 검색 실패 ({folder}): {exc}")
            continue
        if not eids:
            continue
        for eid in eids:
            try:
                _, msg_data = mail.fetch(eid, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                subject = decode_str(msg.get("Subject", ""))
            except Exception as exc:
                print(f"HF 메일 FETCH 실패 ({eid}): {exc}")
                continue

            plain, html = "", ""
            for part in msg.walk():
                ctype = part.get_content_type()
                if header_str(part.get("Content-Disposition")).lower().count("attachment"):
                    continue
                try:
                    payload = part.get_payload(decode=True) or b""
                    cs = part.get_content_charset() or "utf-8"
                    text = payload.decode(cs, errors="replace")
                except Exception:
                    continue
                if ctype == "text/plain":
                    plain += text + "\n"
                elif ctype == "text/html":
                    html += text + "\n"
            body_text = plain or strip_html(html)
            if not body_text:
                continue
            parsed = parse_hf_loan_notice(body_text)
            if not parsed:
                continue

            # lazy load 보금자리론 워크시트
            if loan_ws is None:
                try:
                    import gspread
                    sh = _open_workbook_backend()
                    try:
                        loan_ws = sh.worksheet("보금자리론")
                    except gspread.WorksheetNotFound:
                        loan_ws = sh.add_worksheet("보금자리론", rows=240, cols=8)
                        loan_ws.append_row(
                            ["납입일", "회차", "납입액", "원금", "이자", "잔액", "원본"]
                        )
                    # 기존 회차 수집
                    for r in loan_ws.get_all_values()[1:]:
                        if len(r) >= 2 and r[1]:
                            existing_rounds.add(r[1])
                except Exception as exc:
                    print(f"보금자리론 워크시트 접근 실패: {exc}")
                    return added

            if str(parsed["회차"]) in existing_rounds:
                print(f"  · HF {parsed['회차']}회차 — 이미 등록됨, skip")
                continue
            loan_ws.append_row([
                parsed["납입일"], parsed["회차"], parsed["납입액"],
                parsed["원금"], parsed["이자"], parsed["잔액"],
                body_text[:300],
            ], value_input_option="USER_ENTERED")
            existing_rounds.add(str(parsed["회차"]))
            added += 1
            print(f"  · HF {parsed['회차']}회차 자동 추가 (잔액 {parsed['잔액']:,})")

            if ENABLE_EMAIL_CLEANUP and "처리완료" in dest_folders:
                move_email(mail, eid, dest_folders["처리완료"])

        if ENABLE_EMAIL_CLEANUP:
            try:
                mail.expunge()
            except Exception:
                pass

    return added


# ── Google Sheets 저장 ────────────────────────────────
def _open_workbook_backend():
    """저장 백엔드 선택: STORAGE=sqlite면 로컬 SQLite, 아니면 Google Sheets.

    로컬 모드에서는 PC의 data/budget.db 파일에 직접 기록 —
    localdb.LocalWorkbook이 gspread API 표면을 그대로 제공하므로
    호출부(save_to_sheets, 보금자리론)는 백엔드를 구분하지 않는다.
    """
    if os.environ.get("STORAGE", "").lower() == "sqlite":
        import sys
        repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if repo_root not in sys.path:
            sys.path.insert(0, repo_root)
        from localdb import open_workbook
        return open_workbook(
            os.environ.get("DB_PATH", os.path.join(repo_root, "data", "budget.db"))
        )

    import json
    import tempfile
    import gspread
    from google.oauth2.service_account import Credentials

    creds_dict = json.loads(GOOGLE_CREDS_JSON)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        json.dump(creds_dict, f)
        creds_path = f.name
    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    os.unlink(creds_path)
    return gc.open_by_key(GOOGLE_SHEET_ID)


def save_to_sheets(transactions: list):
    import gspread

    sheet = _open_workbook_backend()
    print(f"📊 저장소: {sheet.title} → {sheet.url}")

    if not transactions:
        print("새로운 거래 없음")
        return 0

    # 시트 스키마: 날짜·시간·출처·유형·금액·내역·카테고리·원문·잔액·입력경로 (10열)
    # 대시보드(app.py)와 동일한 컬럼 순서로 써야 컬럼이 어긋나지 않음.
    SHEET_COLS = ["날짜", "시간", "출처", "유형", "금액", "내역",
                  "카테고리", "원문", "잔액", "입력경로"]
    try:
        ws = sheet.worksheet("거래내역")
    except gspread.WorksheetNotFound:
        ws = sheet.add_worksheet("거래내역", rows=10000, cols=12)
        ws.append_row(SHEET_COLS)

    existing = ws.get_all_values()
    existing_keys = set()
    for row in existing[1:]:
        if len(row) >= 6:
            # 날짜_출처_금액_내역
            existing_keys.add(f"{row[0]}_{row[2]}_{row[4]}_{row[5]}")

    new_rows = []
    for tx in transactions:
        key = f"{tx['날짜']}_{tx['출처']}_{tx['금액']}_{tx['내역']}"
        if key in existing_keys:
            continue
        existing_keys.add(key)
        # email_parser는 cron에서만 실행 → 전부 자동 수집. tx가 명시 안 하면 "자동:{출처}".
        path = tx.get("입력경로") or f"자동:{tx['출처']}"
        new_rows.append([
            tx["날짜"], tx["시간"], tx["출처"], tx["유형"],
            tx["금액"], tx["내역"], tx["카테고리"], tx["원문"],
            tx.get("잔액", ""), path,
        ])

    if new_rows:
        ws.append_rows(new_rows, value_input_option="USER_ENTERED")
        print(f"✅ {len(new_rows)}개 거래 저장 완료")
    else:
        print("중복 없음, 새 거래 없음")
    return len(new_rows)


def build_summary_email(saved_count: int, total_moved: dict,
                        parsed_count: int = 0) -> tuple:
    """수집 결과 발신용 (제목, 본문). 알릴 게 없으면 (None, None).

    '조용한 성공-실패' 방지가 목적 — 명세서·카뱅 내보내기 같은 월 1회
    이벤트가 처리됐을 때만 보내므로 월 몇 통 수준. 매일 빈 결과까지
    보내 소음이 되지 않게 저장 0건 + 이벤트 0건이면 안 보낸다.
    """
    events = {k: v for k, v in (total_moved or {}).items()
              if v and k in ("명세서", "KB명세서", "카뱅내보내기", "보금자리론")}
    if saved_count <= 0 and not events:
        return None, None

    lines = [f"저장된 거래: {saved_count}건 (파싱 {parsed_count}건, 중복 제외 후)"]
    label = {"명세서": "BC카드 명세서", "KB명세서": "KB카드 명세서",
             "카뱅내보내기": "카카오뱅크 내보내기", "보금자리론": "보금자리론 회차"}
    for k, v in events.items():
        lines.append(f"· {label.get(k, k)}: {v}건 처리")
    if parsed_count > 0 and saved_count == 0:
        lines.append("⚠️ 파싱은 됐지만 전부 중복 — 이미 시트에 있는 데이터")
    lines.append("")
    lines.append("대시보드에서 확인 후 카테고리가 '기타'인 거래를 정리해 주세요.")

    subject = f"[가계부] 수집 결과: {saved_count}건 저장 ({datetime.now():%m/%d})"
    return subject, "\n".join(lines)


def send_summary_email(subject: str, body: str) -> bool:
    """네이버 SMTP로 요약 발신 (수신: 본인, SUMMARY_EMAIL_TO로 변경 가능).

    IMAP과 같은 NAVER_EMAIL/NAVER_APP_PW를 쓰므로 추가 시크릿 불필요.
    발신 실패는 수집을 깨면 안 되므로 호출부에서 예외를 삼킨다.
    """
    import smtplib
    from email.mime.text import MIMEText

    to_addr = os.environ.get("SUMMARY_EMAIL_TO", "").strip() or NAVER_EMAIL
    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = NAVER_EMAIL
    msg["To"] = to_addr
    with smtplib.SMTP("smtp.naver.com", 587, timeout=30) as smtp:
        smtp.starttls()
        smtp.login(NAVER_EMAIL, NAVER_APP_PW)
        smtp.send_message(msg)
    print(f"📧 요약 메일 발신 → {to_addr}")
    return True


def build_cleanup_summary(cleanup_log: dict, total_moved: dict) -> str:
    """이메일 정리 결과를 사람이 읽기 좋은 요약 텍스트로.

    카테고리별 처리 건수 + 대표 제목 몇 개를 보여준다.
    홍보(광고) 메일은 읽음 처리되어 정리됨을 명시.
    """
    icons = {"쇼핑": "🛒", "SNS": "📱", "뉴스레터": "📰", "광고": "📢"}
    lines = ["", "─" * 40, "📬 이메일 정리 요약"]
    any_cleaned = False
    for cat in ("광고", "쇼핑", "SNS", "뉴스레터"):
        subjects = cleanup_log.get(cat, [])
        n = total_moved.get(cat, 0)
        if not n:
            continue
        any_cleaned = True
        icon = icons.get(cat, "•")
        read_note = " (읽음 처리)" if cat == "광고" else ""
        lines.append(f"{icon} {cat}: {n}건{read_note}")
        for subj in subjects[:3]:
            lines.append(f"    - {subj}")
        if len(subjects) > 3:
            lines.append(f"    ... 외 {len(subjects) - 3}건")
    if not any_cleaned:
        lines.append("정리할 비거래 메일 없음")
    # 거래/명세서 처리 요약
    tx_moved = total_moved.get("거래", 0)
    stmt = total_moved.get("명세서", 0) + total_moved.get("KB명세서", 0)
    if tx_moved or stmt:
        lines.append(f"💳 거래 알림 {tx_moved}건 + 명세서 {stmt}건 → 처리완료 폴더")
    lines.append("─" * 40)
    return "\n".join(lines)


# ── 메인 ──────────────────────────────────────────────
def main():
    print(f"[{datetime.now()}] 이메일 파싱 시작 (cleanup={'on' if ENABLE_EMAIL_CLEANUP else 'off'})...")
    mail = connect_imap()

    dest_folders = prepare_dest_folders(mail)

    all_transactions = []
    total_moved = {"거래": 0, "쇼핑": 0, "SNS": 0, "뉴스레터": 0, "광고": 0}
    cleanup_log = {}  # {카테고리: [제목,...]} — 정리 요약용

    for folder in IMAP_FOLDERS:
        try:
            transactions, moved = process_folder(
                mail, folder, LOOKBACK_HOURS, dest_folders, cleanup_log
            )
            all_transactions.extend(transactions)
            for k, v in moved.items():
                total_moved[k] = total_moved.get(k, 0) + v
        except Exception as exc:
            print(f"폴더 처리 실패 ({folder}): {exc}")

    # BC카드 / IBK BC카드 월간 명세서 PDF 별도 패스 (LOOKBACK_DAYS 윈도우)
    # IBK BC카드는 비암호 PDF라 BC_PDF_PASSWORD 없어도 처리 가능 — 항상 시도
    try:
        stmt_txs, stmt_moved = process_statements(mail, IMAP_FOLDERS, dest_folders)
        all_transactions.extend(stmt_txs)
        total_moved["명세서"] = stmt_moved
    except Exception as exc:
        print(f"명세서 처리 실패: {exc}")
    if not BC_PDF_PASSWORD:
        print("ℹ️  BC_PDF_PASSWORD 미설정: BC카드 본사 암호 PDF는 건너뜀 (IBK BC카드 비암호 PDF는 처리)")

    # KB국민카드 이메일 명세서 HTML 별도 패스
    try:
        kb_txs, kb_moved = process_kb_statements(mail, IMAP_FOLDERS, dest_folders)
        all_transactions.extend(kb_txs)
        total_moved["KB명세서"] = kb_moved
    except Exception as exc:
        print(f"KB 명세서 처리 실패: {exc}")

    # 카카오뱅크 '거래내역 엑셀' 내보내기 별도 패스
    # 사용자가 앱에서 '내보내기 → 이메일'만 누르면 첨부를 자동 수집.
    try:
        kk_txs, kk_moved = process_kakao_exports(mail, IMAP_FOLDERS, dest_folders)
        all_transactions.extend(kk_txs)
        total_moved["카뱅내보내기"] = kk_moved
    except Exception as exc:
        print(f"카뱅 내보내기 처리 실패: {exc}")

    # 주택금융공사 보금자리론 안내 — 회차 정보를 보금자리론 워크시트에 자동 누적
    try:
        hf_added = process_hf_loan_emails(mail, IMAP_FOLDERS, dest_folders)
        if hf_added:
            print(f"🏠 주택금융공사: {hf_added}개 회차 추가")
            total_moved["보금자리론"] = hf_added
    except Exception as exc:
        print(f"주택금융공사 안내 처리 실패: {exc}")

    print(f"파싱된 거래: {len(all_transactions)}개")
    if ENABLE_EMAIL_CLEANUP:
        moved_summary = ", ".join(f"{k} {v}" for k, v in total_moved.items() if v)
        print(f"이동된 메일: {moved_summary or '없음'}")
        print(build_cleanup_summary(cleanup_log, total_moved))

    try:
        mail.logout()
    except Exception:
        pass

    saved = save_to_sheets(all_transactions) or 0

    # SMTP 요약 발신 — 저장/명세서 이벤트가 있을 때만 (월 몇 통 수준).
    # EMAIL_SUMMARY=false로 끌 수 있고, 발신 실패는 수집 결과에 영향 없음.
    if os.environ.get("EMAIL_SUMMARY", "true").lower() not in ("false", "0", "no"):
        try:
            subject, body = build_summary_email(
                saved, total_moved, parsed_count=len(all_transactions))
            if subject:
                send_summary_email(subject, body)
        except Exception as exc:
            print(f"요약 메일 발신 실패 (수집은 정상): {exc}")

    print("완료!")


if __name__ == "__main__":
    main()
